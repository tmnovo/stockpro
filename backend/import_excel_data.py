"""
Import real data from Excel files into MongoDB.
- Wipes existing clients, products, suppliers, orders, order_items.
- Keeps users and settings untouched.
"""
import os
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from pymongo import MongoClient

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

client = MongoClient(MONGO_URL)
db = client[DB_NAME]

CLIENTS_XLSX = "/app/php_app/data/LISTA_CLIENTES_EXCEL.xlsx"
PRODUCTS_XLSX = "/app/php_app/data/PRODUTOS_EXCEL.xlsx"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def iso_now():
    return datetime.now(timezone.utc).isoformat()


def import_clients():
    wb = openpyxl.load_workbook(CLIENTS_XLSX, data_only=True)
    ws = wb.active
    # Row 7 = header, rows 8+ = data
    # Columns (1-based): 1=Tax region, 2=NIF, 3=Nome, 4=Telemóvel, 5=Telefone,
    # 6=E-mail, 7=S.P., 8=RIC, 9=Sub-conta, 10=Morada, 11=CP, 12=Localidade, 13=País
    records = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        if not row or all(v is None for v in row):
            continue
        nif = clean(row[1])
        name = clean(row[2])
        if not name:
            continue
        mobile = clean(row[3])
        phone = clean(row[4])
        email = clean(row[5])
        address = clean(row[9])
        postal = clean(row[10])
        city = clean(row[11])
        full_addr = ", ".join([p for p in [address, postal, city] if p]) or None

        records.append({
            "id": str(uuid.uuid4()),
            "name": name,
            "email": email,
            "phone": phone or mobile,
            "mobile": mobile,
            "address": full_addr,
            "tax_id": nif,
            "city": city,
            "postal_code": postal,
            "notes": None,
            "created_at": iso_now(),
            "updated_at": iso_now(),
        })

    db.clients.delete_many({})
    if records:
        db.clients.insert_many(records)
    return len(records)


def import_products():
    wb = openpyxl.load_workbook(PRODUCTS_XLSX, data_only=True)
    ws = wb.active
    # Row 8 = header, rows 9+ = data
    # Columns: 1=Tipo, 2=Código, 3=EAN, 4=Família, 5=Descrição/Nome, 6=Uni.,
    # 7=Preço sem IVA, 8=Preço venda, 9=IVA incluído, 10=Preço compra, 11=Taxa IVA
    records = []
    families = set()
    for row in ws.iter_rows(min_row=9, values_only=True):
        if not row or all(v is None for v in row):
            continue
        tipo = clean(row[0])
        code = clean(row[1])
        ean = clean(row[2])
        familia = clean(row[3])
        name = clean(row[4])
        unit = clean(row[5]) or "un"
        price_no_vat = row[6]
        price_sell = row[7]

        if not name:
            continue
        # Use sell price if present, otherwise no-vat price
        try:
            price = float(price_sell) if price_sell not in (None, "") else float(price_no_vat or 0)
        except (TypeError, ValueError):
            price = 0.0

        if familia:
            families.add(familia)

        records.append({
            "id": str(uuid.uuid4()),
            "name": name,
            "sku": code,
            "barcode": ean,
            "description": familia,
            "category": familia,
            "type": tipo,
            "price": round(price, 4),
            "stock": 0,
            "unit": unit,
            "created_at": iso_now(),
            "updated_at": iso_now(),
        })

    db.products.delete_many({})
    if records:
        db.products.insert_many(records)
    return len(records), sorted(families)


def wipe_orders():
    db.orders.delete_many({})
    db.order_items.delete_many({})


def main():
    print(">> Wiping existing orders…")
    wipe_orders()

    print(">> Importing clients from", CLIENTS_XLSX)
    nc = import_clients()
    print(f"   Imported {nc} clients")

    print(">> Importing products from", PRODUCTS_XLSX)
    np, fams = import_products()
    print(f"   Imported {np} products in {len(fams)} families")

    print("\n=== Final counts ===")
    print("Clients :", db.clients.count_documents({}))
    print("Products:", db.products.count_documents({}))
    print("Orders  :", db.orders.count_documents({}))
    print("Users   :", db.users.count_documents({}))


if __name__ == "__main__":
    main()
